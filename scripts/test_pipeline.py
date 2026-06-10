"""Test du pipeline 3 couches (Generator → Judge → Refine) sur l'Excel de questions.

Prend N questions de CHAQUE feuille du classeur (2 par défaut), exécute le pipeline
complet sur le cours choisi, et exporte un Excel récapitulatif dans data_eval/ :
id, feuille, question, reponse_attendue, reponse_finale, verdict, iterations,
methode, reponse_v1.

Usage :
    python scripts/test_pipeline.py                       # cours par défaut, 2 q/feuille
    python scripts/test_pipeline.py --cours ch11_pgcd_diviseurs --par-feuille 2
    python scripts/test_pipeline.py --questions data/Lot2_questions_CH11.xlsx

⚠️ Consomme du quota OpenRouter (free tier) : ~1 à 5 appels par question. Le script
logge la progression et continue malgré les erreurs ponctuelles.
"""

import argparse
import logging
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Make the project root importable when run as a script.
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.config import resolve_api_key, charge_cours, liste_cours
from core.latex import nettoie_latex
from core.pipeline.orchestrator import repond

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("test_pipeline")

# Verdict → header fill colour for the output sheet.
_FILLS = {
    "OK":          "C6EFCE",   # green
    "NOK":         "FFEB9C",   # orange
    "CALCUL_FAUX": "FFC7CE",   # red
    "ERREUR":      "D9D9D9",   # grey
}
COLONNES = ["id", "feuille", "question", "reponse_attendue",
            "reponse_finale", "verdict", "iterations", "methode", "reponse_v1"]


def lire_questions(xlsx_path: str, par_feuille: int) -> list[dict]:
    """Lit les `par_feuille` premières questions de chaque feuille du classeur."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    items = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        try:
            i_id = header.index("id")
            i_q = header.index("question")
            i_rep = header.index("reponse_attendue")
        except ValueError:
            logger.warning("Feuille %s ignorée (colonnes inattendues)", ws.title)
            continue
        pris = 0
        for row in rows[1:]:
            if pris >= par_feuille:
                break
            if row[i_q] is None:
                continue
            items.append({
                "id": row[i_id], "feuille": ws.title,
                "question": str(row[i_q]),
                "reponse_attendue": str(row[i_rep]) if row[i_rep] is not None else "",
            })
            pris += 1
    return items


def exporte(resultats: list[dict], out_path: str) -> None:
    """Écrit le récapitulatif Excel stylé (en-tête coloré, colonnes lisibles)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    for col, nom in enumerate(COLONNES, start=1):
        c = ws.cell(row=1, column=col, value=nom)
        c.font = header_font
        c.fill = header_fill

    for r, res in enumerate(resultats, start=2):
        for col, nom in enumerate(COLONNES, start=1):
            cell = ws.cell(row=r, column=col, value=res.get(nom, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Colour the verdict cell.
        verdict = res.get("verdict", "")
        if verdict in _FILLS:
            i_verdict = COLONNES.index("verdict") + 1
            ws.cell(row=r, column=i_verdict).fill = PatternFill("solid", fgColor=_FILLS[verdict])

    widths = {"question": 40, "reponse_attendue": 40, "reponse_finale": 40, "reponse_v1": 40}
    for col, nom in enumerate(COLONNES, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = widths.get(nom, 16)
    ws.freeze_panes = "A2"

    wb.save(out_path)


def main() -> int:
    cours_dispo = liste_cours()
    p = argparse.ArgumentParser(description="Test du pipeline 3 couches.")
    p.add_argument("--cours", default=cours_dispo[0] if cours_dispo else None,
                   help=f"Slug du cours (dispo : {cours_dispo})")
    p.add_argument("--questions", default="data/Lot2_questions_CH11.xlsx",
                   help="Classeur Excel des questions.")
    p.add_argument("--par-feuille", type=int, default=2,
                   help="Nombre de questions à prendre par feuille (défaut 2).")
    p.add_argument("--out", default=None, help="Chemin du fichier de sortie .xlsx.")
    args = p.parse_args()

    if not args.cours:
        logger.error("Aucun cours trouvé dans cours/. Crée d'abord un dossier de cours.")
        return 1

    api_key = resolve_api_key()
    if not api_key:
        logger.error("OPENROUTER_API_KEY introuvable (env ou .streamlit/secrets.toml).")
        return 1

    cours_brut, verif_path = charge_cours(args.cours)
    cours_texte = nettoie_latex(cours_brut)
    logger.info("Cours « %s » chargé (%d car.) — verif : %s",
                args.cours, len(cours_texte), verif_path or "aucun")

    items = lire_questions(args.questions, args.par_feuille)
    logger.info("%d questions à traiter (%d/feuille)\n", len(items), args.par_feuille)

    resultats = []
    for n, item in enumerate(items, start=1):
        logger.info("─── [%d/%d] %s · q%s : %.60s",
                    n, len(items), item["feuille"], item["id"], item["question"])
        try:
            out = repond(item["question"], cours_texte=cours_texte,
                        api_key=api_key, verif_path=verif_path)
        except Exception as e:   # never let one question abort the whole run
            logger.error("    ÉCHEC inattendu : %s", e)
            out = {"reponse": "", "verdict": "ERREUR", "iterations": 0,
                   "methode": "exception", "reponse_v1": ""}
        logger.info("    → %s | iter %d | %s\n",
                    out["verdict"], out["iterations"], out["methode"])
        resultats.append({
            **item,
            "reponse_finale": out["reponse"],
            "verdict": out["verdict"],
            "iterations": out["iterations"],
            "methode": out["methode"],
            "reponse_v1": out["reponse_v1"],
        })

    out_path = args.out or f"data_eval/test_pipeline_{args.cours}.xlsx"
    exporte(resultats, out_path)
    logger.info("✓ Export : %s (%d lignes)", out_path, len(resultats))

    # Quick verdict tally.
    tally: dict[str, int] = {}
    for r in resultats:
        tally[r["verdict"]] = tally.get(r["verdict"], 0) + 1
    logger.info("Récap verdicts : %s", tally)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
